import csv, re, os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE        = r"C:\Users\azaol\OneDrive\Desktop\Figure-Connect-CMF Stuff"
DOC1_PATH   = os.path.join(BASE, "CMH_Fig-Connect_Mapping Specs_20260629.xlsx")
DOC2_PATH   = os.path.join(BASE, "mssql_Field_Mapping_nopii_Diff_6_22_26.xlsx")
FIG_BASE    = os.path.join(BASE, "Figure Connect csv Schemas")
OUT_PATH    = r"C:\Users\azaol\OneDrive\Desktop\CMH_Figure_Mapping_Output_20260629_v2.xlsx"

ABBR = {
    'borr':'borrower','bwr':'borrower','bor':'borrower',
    'amt':'amount','pmt':'payment','pymt':'payment',
    'prop':'property','addr':'address',
    'mo':'monthly','mthly':'monthly',
    'int':'interest','orig':'origination',
    'uw':'underwriting','lo':'loanofficer',
    'nbr':'number','num':'number',
    'dt':'date','exp':'expiration',
    'val':'value','comp':'compensation',
    'cx':'custom','acq':'acquisition',
}

def tokenize(s):
    if not s: return set()
    toks = re.findall(r'[a-z]+', s.lower())
    return {ABBR.get(t, t) for t in toks}

# ── manual Figure mappings ────────────────────────────────────────────────────
M = {
    'CX_FUNDSORDWIREAMT':[('warehouse-funding-queue','draw_amount','PRIMARY','Wire/funds amount = draw_amount')],
    'F2':  [('warehouse-funding-queue','application_amount','PRIMARY','Total loan amount'),('application-event','app_amount','SECONDARY','')],
    'TotalLoanAmount':[('warehouse-funding-queue','application_amount','PRIMARY',''),('application-event','app_amount','SECONDARY','')],
    'F3':  [('warehouse-funding-queue','interest_rate','PRIMARY',''),('application-event','app_rate','SECONDARY','')],
    'InterestRate':[('warehouse-funding-queue','interest_rate','PRIMARY',''),('application-event','app_rate','SECONDARY','')],
    'F4':  [('warehouse-funding-queue','loan_term','PRIMARY','Term in years; multiply by 12 for payment count'),('application-event','app_term','SECONDARY','')],
    'NumberofPayments':[('warehouse-funding-queue','loan_term','PRIMARY','Figure stores term in years; payments = term x 12')],
    'F5':  [('warehouse-funding-queue','monthly_payment','PRIMARY',''),('application-event','selected_offer_estimated_monthly_payment','SECONDARY','')],
    'PIPayment':[('warehouse-funding-queue','monthly_payment','PRIMARY','')],
    'TotalProposedMonthlyPaymentForProperty':[('warehouse-funding-queue','monthly_payment','PRIMARY','')],
    'F11': [('warehouse-funding-queue','property_street_address','PRIMARY',''),('application-event','prop_street_address','SECONDARY','')],
    'PropertyAddress':[('warehouse-funding-queue','property_street_address','PRIMARY',''),('application-event','prop_street_address','SECONDARY','')],
    'F12': [('warehouse-funding-queue','prop_city','PRIMARY',''),('application-event','prop_city','SECONDARY','')],
    'PropertyCity':[('warehouse-funding-queue','prop_city','PRIMARY',''),('application-event','prop_city','SECONDARY','')],
    'F13': [('application-event','property_county','PRIMARY','Not in warehouse-funding-queue; use application-event')],
    'PropertyCounty':[('application-event','property_county','PRIMARY','')],
    'F14': [('warehouse-funding-queue','prop_state','PRIMARY',''),('application-event','prop_state','SECONDARY','')],
    'PropertyState':[('warehouse-funding-queue','prop_state','PRIMARY',''),('application-event','prop_state','SECONDARY','')],
    'F15': [('warehouse-funding-queue','prop_zipcode','PRIMARY',''),('application-event','prop_zip','SECONDARY','')],
    'PropertyZip':[('warehouse-funding-queue','prop_zipcode','PRIMARY',''),('application-event','prop_zip','SECONDARY','')],
    'F136':[('application-event','avm_value_amount','PRIMARY','HELOC FLAG: No sales price in HELOC; AVM value is closest equivalent.')],
    'SalesPrice':[('application-event','avm_value_amount','PRIMARY','HELOC FLAG: No sales price concept in HELOC; AVM property value is closest semantic match.')],
    'F142':[('warehouse-funding-queue','draw_amount','PRIMARY','Cash to borrower = draw_amount'),('application-event','draw_amount','SECONDARY','')],
    'CashFromBorrower':[('warehouse-funding-queue','draw_amount','PRIMARY','')],
    'F247':[None],'LifeCap':[None],
    'F325':[None],'BalloonTerm':[None],
    'F356':[('warehouse-funding-queue','property_adjusted_value','PRIMARY',''),('application-event','avm_value_amount','SECONDARY','')],
    'UWApprovalMonitoringAppraisalValue':[('warehouse-funding-queue','property_adjusted_value','PRIMARY',''),('application-event','avm_value_amount','SECONDARY','')],
    'F364':[('warehouse-funding-queue','application_short_id','PRIMARY','Figure application_short_id = loan number equivalent'),('application-event','application_short_id','SECONDARY','')],
    'Loannumber':[('warehouse-funding-queue','application_short_id','PRIMARY',''),('application-event','application_short_id','SECONDARY','')],
    'LoanNumber':[('warehouse-funding-queue','application_short_id','PRIMARY',''),('application-event','application_short_id','SECONDARY','')],
    'F37': [('warehouse-funding-queue','borrower_last_name','PRIMARY','Figure has no suffix field'),('application-event','borrower_1_last_name','SECONDARY','')],
    'LastNameWithSuffix':[('warehouse-funding-queue','borrower_last_name','PRIMARY','Figure does not store suffix separately')],
    'F420':[('warehouse-funding-queue','lien_position','PRIMARY',''),('application-event','figure_new_lien_position','SECONDARY','')],
    'LienPriority':[('warehouse-funding-queue','lien_position','PRIMARY',''),('application-event','figure_new_lien_position','SECONDARY','')],
    'F428':[('application-event','lien_position_2_current_balance','PRIMARY','Second mortgage balance = lien position 2 current balance')],
    'SecondMtgLoanAmount':[('application-event','lien_position_2_current_balance','PRIMARY','')],
    'F608':[('warehouse-funding-queue','rate_type','PRIMARY','HELOC FLAG: Amortization type maps to rate_type (FIXED/VARIABLE)'),('application-event','rate_type','SECONDARY','')],
    'AmortizationTypeEnum':[('warehouse-funding-queue','rate_type','PRIMARY','HELOC FLAG: Figure values are FIXED or VARIABLE'),('application-event','rate_type','SECONDARY','')],
    'F682':[('application-event','first_payment_due_date','PRIMARY','')],
    'DateFirstPaymentDue':[('application-event','first_payment_due_date','PRIMARY','')],
    'F689':[None],'ARMMargin':[None],
    'F696':[('warehouse-funding-queue','draw_term','PRIMARY','HELOC FLAG: draw_term is HELOC draw period in years')],
    'InitialFixedRatePeriod':[('warehouse-funding-queue','draw_term','PRIMARY','HELOC FLAG: Nearest equivalent is HELOC draw_term; semantics differ')],
    'F697':[None],'PeriodicRateCap':[None],
    'F736':[('warehouse-funding-queue','verified_income_amount','PRIMARY',''),('application-event','monthly_income','SECONDARY','Note: application-event is monthly')],
    'TotalIncomeAmount':[('warehouse-funding-queue','verified_income_amount','PRIMARY',''),('application-event','monthly_income','SECONDARY','')],
    'F740':[('warehouse-funding-queue','post_loan_dti','PRIMARY','HELOC FLAG: Figure combines housing+total debt into single DTI'),('application-event','post_loan_dti','SECONDARY','')],
    'HousingRatio':[('warehouse-funding-queue','post_loan_dti','PRIMARY','HELOC FLAG: Figure has no separate housing ratio')],
    'F742':[('warehouse-funding-queue','post_loan_dti','PRIMARY',''),('application-event','post_loan_dti','SECONDARY','')],
    'BottomRatioPercent':[('warehouse-funding-queue','post_loan_dti','PRIMARY','')],
    'F748':[('warehouse-funding-queue','heloc_agreement_signed_date','PRIMARY',''),('application-event','Funding_Date','SECONDARY','')],
    'DateClose':[('warehouse-funding-queue','heloc_agreement_signed_date','PRIMARY',''),('application-event','Funding_Date','SECONDARY','')],
    'F912':[('warehouse-funding-queue','monthly_payment','PRIMARY','')],
    'ProposedTotalHousingExpense':[('warehouse-funding-queue','monthly_payment','PRIMARY','')],
    'F1051':[None],'MERSNumber':[None],
    'F1172':[('hmda','LoanType','PRIMARY','HELOC FLAG: Figure is always HELOC/open-end LOC'),('application-event','rate_type','SECONDARY','')],
    'LoanTypeEnum':[('hmda','LoanType','PRIMARY','HELOC FLAG: Figure product is always HELOC; verify mapping')],
    'F186':[None],'EscrowCaseNumber':[None],
    'F187':[None],'TitleCaseNumber':[None],
    'F19': [('application-event','application_reason','PRIMARY',''),('hmda','Purpose','SECONDARY','')],
    'LoanPurposeEnum':[('application-event','application_reason','PRIMARY',''),('hmda','Purpose','SECONDARY','')],
    'F1401':[None],'LoanProgram':[None],
    'F1402':[('warehouse-funding-queue','borrower_date_of_birth','PRIMARY',''),('application-event','borrower_1_dob','SECONDARY','')],
    'BorrBirthdate':[('warehouse-funding-queue','borrower_date_of_birth','PRIMARY',''),('application-event','borrower_1_dob','SECONDARY','')],
    'F1402_P2':[('application-event','borrower_2_dob','PRIMARY','Co-borrower DOB; null if no co-borrower')],
    'F1543':[('application-event','uw_decision_decision','PRIMARY',''),('application-event','uw_decision_sub_decision','SECONDARY','')],
    'UWRiskAssessTypeEnum':[('application-event','uw_decision_decision','PRIMARY',''),('application-event','uw_decision_sub_decision','SECONDARY','')],
    'F1544':[('application-event','uw_decision_sub_decision','PRIMARY','F1544=AUSRecommendation (Callidus SQL confirmed); no AUS in Figure HELOC; UW sub-decision is closest equivalent')],
    'F1553':[('application-event','property_type','PRIMARY',''),('warehouse-funding-queue','property_land_use','SECONDARY','')],
    'SubjectPropertyTypeEnum':[('application-event','property_type','PRIMARY',''),('warehouse-funding-queue','property_land_use','SECONDARY','')],
    'F1742':[('warehouse-funding-queue','monthly_payment','PRIMARY',''),('application-event','selected_offer_estimated_monthly_payment','SECONDARY','')],
    'F1811':[('warehouse-funding-queue','occupancy_type','PRIMARY','HELOC FLAG: Figure values = HOME/INVESTMENT/SECONDARY')],
    'OccupancyStatusEnum':[('warehouse-funding-queue','occupancy_type','PRIMARY','HELOC FLAG: Figure uses HOME/INVESTMENT/SECONDARY; map to CMH enum')],
    'F1821':[('warehouse-funding-queue','property_adjusted_value','PRIMARY',''),('application-event','avm_value_amount','SECONDARY','')],
    'PropertyValue':[('warehouse-funding-queue','property_adjusted_value','PRIMARY',''),('application-event','avm_value_amount','SECONDARY','')],
    'F1996':[('warehouse-funding-queue','funds_transfer_effective_date','PRIMARY',''),('funding-transaction','batch_effective_date','SECONDARY','')],
    'DateFundsOrdered':[('warehouse-funding-queue','funds_transfer_effective_date','PRIMARY',''),('funding-transaction','batch_effective_date','SECONDARY','')],
    'F2001':[None],'FundsWireTo':[None],
    'F2358':[('application-event','bpo_type','PRIMARY','BPO type is Figure equivalent of appraisal type'),('application-event','property_valuation_source_category','SECONDARY','')],
    'AppraisalTypeEnum':[('application-event','bpo_type','PRIMARY',''),('application-event','property_valuation_source_category','SECONDARY','')],
    'F2626':[None],'LoanChannelEnum':[None],
    'F4000':[('warehouse-funding-queue','borrower_first_name','PRIMARY',''),('application-event','borrower_1_first_name','SECONDARY','')],
    'BorrFirst':[('warehouse-funding-queue','borrower_first_name','PRIMARY',''),('application-event','borrower_1_first_name','SECONDARY','')],
    'F4002':[('warehouse-funding-queue','borrower_last_name','PRIMARY',''),('application-event','borrower_1_last_name','SECONDARY','')],
    'BorrLast':[('warehouse-funding-queue','borrower_last_name','PRIMARY',''),('application-event','borrower_1_last_name','SECONDARY','')],
    'F4002_P2':[('application-event','borrower_2_last_name','PRIMARY','Co-borrower last name; null if no co-borrower')],
    'F4004':[('warehouse-funding-queue','borrower_first_name','PRIMARY','Duplicate of F4000'),('application-event','borrower_1_first_name','SECONDARY','')],
    'F4004_P2':[('application-event','borrower_2_first_name','PRIMARY','Co-borrower first name; null if no co-borrower')],
    'F4006':[('warehouse-funding-queue','borrower_last_name','PRIMARY',''),('application-event','borrower_1_last_name','SECONDARY','')],
    'MORNET_X4':[None],'AUSKey':[None],
    'MORNET_X67':[None],'DocTypeEnum':[None],
    'VASUMM_X23':[('warehouse-funding-queue','credit_hard_fico','PRIMARY',''),('application-event','credit_hard_fico','SECONDARY','')],
    'UWApprvalMonitoringFICO':[('warehouse-funding-queue','credit_hard_fico','PRIMARY',''),('application-event','credit_hard_fico','SECONDARY','')],
    'VEND_X396':[None],'EscrowCoABANbr':[None],
    'VEND_X397':[None],'VEND_X399':[None],
    'VEND_X398':[None],'TitleCoABANbr':[None],
    'HMDA_X13':[('hmda','Action','PRIMARY','HMDA action taken field')],
    # HMDA race/ethnicity indicators (F1523-F1538) — reference hmda.csv; ETL required
    'F1523':[('hmda','Ethnicity','PRIMARY','HMDA ethnicity indicator; reference hmda.csv schema; ETL required: CMH binary indicators -> Figure categorical coded values')],
    'F1524':[('hmda','Ethnicity','PRIMARY','HMDA ethnicity - type detail; reference hmda.csv; ETL required')],
    'F1525':[('hmda','Ethnicity','PRIMARY','HMDA ethnicity - other detail; reference hmda.csv; ETL required')],
    'F1526':[('hmda','Ethnicity','PRIMARY','HMDA co-applicant ethnicity; reference hmda.csv; ETL required')],
    'F1527':[('hmda','Ethnicity','PRIMARY','HMDA co-applicant ethnicity detail; reference hmda.csv; ETL required')],
    'F1528':[('hmda','Race','PRIMARY','HMDA race indicator; reference hmda.csv; ETL required')],
    'F1529':[('hmda','Race','PRIMARY','HMDA race - American Indian detail; reference hmda.csv; ETL required')],
    'F1530':[('hmda','Race','PRIMARY','HMDA race - Asian detail; reference hmda.csv; ETL required')],
    'F1531':[('hmda','Race','PRIMARY','HMDA race - NHPI detail; reference hmda.csv; ETL required')],
    'F1532':[('hmda','Race','PRIMARY','HMDA race - other detail; reference hmda.csv; ETL required')],
    'F1533':[('hmda','Race','PRIMARY','HMDA co-applicant race; reference hmda.csv; ETL required')],
    'F1534':[('hmda','Race','PRIMARY','HMDA co-applicant race detail; reference hmda.csv; ETL required')],
    'F1535':[('hmda','Race','PRIMARY','HMDA co-applicant race Asian detail; reference hmda.csv; ETL required')],
    'F1536':[('hmda','Race','PRIMARY','HMDA co-applicant race other detail; reference hmda.csv; ETL required')],
    'F1537':[('hmda','Race','PRIMARY','HMDA race - additional detail; reference hmda.csv; ETL required')],
    'F1538':[('hmda','Race','PRIMARY','HMDA race - co-applicant additional; reference hmda.csv; ETL required')],
    'F471': [('hmda','Sex','PRIMARY','HMDA sex/gender indicator; reference hmda.csv; ETL required')],
    'F471_P2':[('hmda','Sex','PRIMARY','HMDA co-applicant sex/gender; reference hmda.csv; ETL required')],
    'F478': [('hmda','Sex','PRIMARY','HMDA sex indicator variant; reference hmda.csv; ETL required')],
    # Callidus fields
    'ApprovalStatus':[('application-event','app_status','PRIMARY',''),('application-event','uw_decision_decision','SECONDARY','')],
    'AUSRecommendation':[('application-event','uw_decision_sub_decision','PRIMARY','No AUS in Figure HELOC; UW sub-decision is closest equivalent; FID = F1544')],
    'NMLS_X8':[('user-identity','officer_nmls_no','PRIMARY',''),('application-event','nmls_mlo_id','SECONDARY','')],
    'BranchManagerNMLS':[('user-identity','officer_nmls_no','PRIMARY','')],
    'F313':[('originator-broker','address_city','PRIMARY','')],
    'BrokerCity':[('originator-broker','address_city','PRIMARY','')],
    'F315':[('originator-broker','broker_name','PRIMARY',''),('application-event','wholesale_originator_dba_name','SECONDARY','')],
    'BrokerLenderName':[('originator-broker','broker_name','PRIMARY',''),('application-event','wholesale_originator_dba_name','SECONDARY','')],
    'BrokerState':[('originator-broker','address_state','PRIMARY','')],
    'F2149':[None],'BuySideLockDate':[None],
    'F984':[None],'CertifyingUnderwriter':[None],
    'ClearToCloseStatus':[('application-event','application_stage','PRIMARY','Map CTC to stage = Funding Decision or Rescission in Figure workflow')],
    'LoanTeamMember_Name_Closer':[None],'CloserName':[None],
    'CorrAccountExecutiveEmail':[None],'CorrAccountExecutiveName':[None],
    'TPO_X88':[('originator-broker','broker_name','PRIMARY','TPO delegated UW context; maps to broker/originator in Figure')],
    'CorrespondentUnderwritingDelegatedEnum':[('originator-broker','broker_name','PRIMARY','')],
    'F3292':[('application-event','application_start_date','PRIMARY',''),('hmda','ApplDate','SECONDARY','')],
    'DateApplication':[('application-event','application_start_date','PRIMARY',''),('hmda','ApplDate','SECONDARY','')],
    'Log_MS_Date_Approval':[('application-event','product_select_ts','PRIMARY','Approval timestamp = product_select_ts in Figure workflow')],
    'DateApprove':[('application-event','product_select_ts','PRIMARY','')],
    'Log_MS_Date_Clear_to_Close':[('application-event','rescission_ts','PRIMARY','CTC maps to rescission stage in Figure workflow')],
    'DateClearToClose':[('application-event','rescission_ts','PRIMARY','')],
    'Log_MS_Date_Completion':[('application-event','complete_ts','PRIMARY',''),('mcr','funding_date','SECONDARY','')],
    'DateCompletion':[('application-event','complete_ts','PRIMARY','')],
    'DENIAL_X69':[('application-event','declined_ts','PRIMARY',''),('hmda','ActionDate','SECONDARY','')],
    'DateCreditDenialActionTaken':[('application-event','declined_ts','PRIMARY',''),('hmda','ActionDate','SECONDARY','')],
    'Log_MS_Date_Decision':[('application-event','uw_review_ts','PRIMARY','')],
    'DateDecision':[('application-event','uw_review_ts','PRIMARY','')],
    'LOG_MS_DATE_DOCS_OUT':[('application-event','sign_prom_note_ts','PRIMARY','Docs out = sign promissory note stage in Figure')],
    'DateDocsOut':[('application-event','sign_prom_note_ts','PRIMARY','')],
    'F3089':[None],'DateEscrowInsuranceExpire':[None],
    'Log_MS_DateTime_Funding':[('application-event','Funding_Date','PRIMARY',''),('warehouse-funding-queue','funds_transfer_effective_date','SECONDARY','')],
    'DateFund':[('application-event','Funding_Date','PRIMARY',''),('warehouse-funding-queue','funds_transfer_effective_date','SECONDARY','')],
    'F3152':[None],'DateGFE':[None],
    'Log_MS_DATE_PROCESSING':[('application-event','app_start_ts','PRIMARY','Processing start = application start timestamp')],
    'DateProcess':[('application-event','app_start_ts','PRIMARY','')],
    'Log_MS_Date_Purchased':[('application-event','Transfer_Date','PRIMARY','Loan purchase = ownership transfer date')],
    'DatePurchase':[('application-event','Transfer_Date','PRIMARY','')],
    'Log_MS_Date_Shipping':[None],'DateShip':[None],
    'Log_MS_DATETIME_SUBMITTAL':[('application-event','application_start_date','PRIMARY','')],
    'DateSubmittal':[('application-event','application_start_date','PRIMARY','')],
    'TPO_X4':[('application-event','application_start_date','PRIMARY','TPO submit date maps to application start')],
    'DateTPOSubmit':[('application-event','application_start_date','PRIMARY','')],
    'DecisionStatus':[('application-event','uw_decision_decision','PRIMARY','')],
    'DocsOutStatus':[('application-event','application_stage','PRIMARY','')],
    'F3659':[None],'FHAStreamlineType':[None],
    'Funder':[None],
    'FundingStatus':[('warehouse-funding-queue','transaction_status','PRIMARY',''),('funding-transaction','batching_status','SECONDARY','')],
    'F3237':[('loan-officer-license','loan_officer_license_number','PRIMARY','')],
    'LenderStateLicenseNbr':[('loan-officer-license','loan_officer_license_number','PRIMARY','')],
    'LOAN_LOANFOLDER':[None],'LoanFolder':[None],
    'LoanTeamMember_Name_Loan_Officer':[('user-identity','full_name','PRIMARY',''),('application-event','mlo_full_name','SECONDARY','')],
    'LoanOfficerName':[('user-identity','full_name','PRIMARY',''),('application-event','mlo_full_name','SECONDARY','')],
    'F362':[None],'LoanProcessorName':[None],'Processor':[None],
    'LOANTEAMMEMBER_NAME_LOAN_PROCESSOR':[None],
    'LoanProcessor2Email':[None],'LoanProcessor3Email':[None],
    'Log_MS_Stage':[('application-event','application_stage','PRIMARY','')],
    'NextExpectedMilestone':[('application-event','application_stage','PRIMARY','')],
    'F3238':[('user-identity','officer_nmls_no','PRIMARY',''),('application-event','nmls_mlo_id','SECONDARY','')],
    'OriginatorNMLS':[('user-identity','officer_nmls_no','PRIMARY',''),('application-event','nmls_mlo_id','SECONDARY','')],
    'F2630':[None],'PurchaseAdviceConfirmedDate':[None],
    'Log_MS_Date_Reconciled':[('mcr','funding_date','PRIMARY','')],
    'ReconciledDate':[('mcr','funding_date','PRIMARY','')],
    'F3000':[None],'SectionActType':[None],
    'SystemOfRecord':[None],
    'Log_MS_DateTime_Approval':[('application-event','product_select_ts','PRIMARY','')],
    'TimeApprovalDate':[('application-event','product_select_ts','PRIMARY','')],
    'TimeClearToCloseDate':[('application-event','rescission_ts','PRIMARY','')],
    'LoanTeamMember_Email_Loan_Officer':[('user-identity','loan_officer_email','PRIMARY','')],
    'EmailLoanOfficer':[('user-identity','loan_officer_email','PRIMARY','')],
    'LoanTeamMember_Email_Closer':[None],'EmailCloser':[None],
    'LoanTeamMember_Email_Discloser':[None],'EmailDiscloser':[None],
    'LoanTeamMember_Email_Funder':[None],'EmailFunder':[None],
    'LoanTeamMember_Email_Loan_Processor':[None],'EmailLoanProcessor':[None],
    'LoanTeamMember_Email_Underwriter':[None],'EmailUnderwriter':[None],
    'Encompassid':[None],'ModifiedDate':[None],
    'SubmittalStatus':[('application-event','app_status','PRIMARY','')],
    'F3944':[None],'F3945':[None],
    # data entry error row
    'TotalLoanAmount as decimal(10,2))':[None],
}

# ── Callidus CDM view map ─────────────────────────────────────────────────────
# keyed by uppercase DRS FID or uppercase CDM column name
CDM_MAP = {
    # F-numbers
    'F2':('dbo.vLoan','TotalLoanAmount'),
    'F3':('dbo.vLoan','InterestRate'),
    'F4':('dbo.vLoan','NumberofPayments'),
    'F19':('dbo.vProperty','LoanPurposeEnum'),
    'F313':('dbo.vContactBrokerLender','BrokerCity'),
    'F315':('dbo.vContactBrokerLender','BrokerLenderName'),
    'F362':('dbo.vContactLoanProcessor','LoanProcessorName'),
    'F364':('dbo.vLoan','LoanNumber'),
    'F420':('dbo.vLoanProductData','LienPriority'),
    'F608':('dbo.vLoan','AmortizationTypeEnum'),
    'F984':('dbo.vContactUnderWriter','CertifyingUnderwriter'),
    'F1172':('dbo.vLoan','LoanTypeEnum'),
    'F1401':('dbo.vLoan','LoanProgram'),
    'F1543':('dbo.vTsum','UWRiskAssessTypeEnum'),
    'F1544':('dbo.vTsum','AUSRecommendation'),
    'F1553':('dbo.vTsum','SubjectPropertyTypeEnum'),
    'F1811':('dbo.vApplication','OccupancyStatusEnum'),
    'F2149':('dbo.vRateLock','BuySideLockDate'),
    'F2626':('dbo.vLoan','LoanChannelEnum'),
    'F2630':('dbo.vRateLock','PurchaseAdviceConfirmedDate'),
    'F3089':('dbo.vFhaVaLoan','DateEscrowInsuranceExpire'),
    'F3152':('dbo.vRegulationZ','DateGFE'),
    'F3237':('dbo.vContactBrokerLender','LenderStateLicenseNbr'),
    'F3238':('dbo.vLoan','OriginatorNMLS'),
    'F3292':('dbo.vLoan','DateApplication'),
    'F3659':('dbo.vHudLoanData','FHAStreamlineType'),
    'F4002':('dbo.vLoanBorrower','BorrLast'),
    # Custom FIDs
    'VASUMM_X23':('dbo.vVaLoanData','UWApprvalMonitoringFICO'),
    'NMLS_X8':('dbo.vLoanProductData','BranchManagerNMLS'),
    'MORNET_X67':('dbo.vLoanProductData','DocTypeEnum'),
    'TPO_X88':('dbo.vTPO','CorrespondentUnderwritingDelegatedEnum'),
    'TPO_X4':('dbo.vTPO','DateTPOSubmit'),
    'DENIAL_X69':('dbo.vStatementCreditDenial','DateCreditDenialActionTaken'),
    'CX_BRANCHNMLS':('dbo.vCustomFieldString','CX_BRANCHNMLS'),
    'CX_CENTRALIZED_1':('dbo.vCustomFieldString','CX_CENTRALIZED_1'),
    'CX_CMSPAIDOFFLN':('dbo.vCustomFieldNumeric','CX_CMSPAIDOFFLN'),
    'CX_FISERVPRIORLOANNUMBER':('dbo.vCustomFieldNumeric','CX_FISERVPRIORLOANNUMBER'),
    'CX_MS_FCI_COUNT':('dbo.vCustomFieldNumeric','CX_MS_FCI_COUNT'),
    'CX_MS_READYFORDOCS_COUNT':('dbo.vCustomFieldNumeric','CX_MS_READYFORDOCS_COUNT'),
    'CX_NY_CEMA':('dbo.vCustomFieldString','CX_NY_CEMA'),
    'CX_PRODCENTER':('dbo.vCustomFieldString','CX_PRODCENTER'),
    'CX_PROJECTEDCLOSING2':('dbo.vCustomFieldDate','CX_PROJECTEDCLOSING2'),
    'CX_REFERRAL_EMAIL':('dbo.vCustomFieldString','CX_REFERRAL_EMAIL'),
    'CX_TX50F2':('dbo.vCustomFieldString','CX_TX50F2'),
    'CX_TXHOMEEQUITY':('dbo.vCustomFieldString','CX_TXHOMEEQUITY'),
    'CX_UNFUNDDATE':('dbo.vCustomFieldDate','CX_UNFUNDDATE'),
    'CX_UW_2NDSIGDATE':('dbo.vCustomFieldDate','CX_UW_2NDSIGDATE'),
    'CX_UW_2NDSIGNAME':('dbo.vCustomFieldString','CX_UW_2NDSIGNAME'),
    'CX_UW_CTCBY':('dbo.vCustomFieldString','CX_UW_CTCBY'),
    'CX_UWQC_CNTCONDFORLP':('dbo.vCustomFieldNumeric','CX_UWQC_CNTCONDFORLP'),
    'CUST01FV':('dbo.vCustomFieldString','CUST01FV'),
    'CUST40FV':('dbo.vCustomFieldString','CUST40FV'),
    'CUST46FV':('dbo.vCustomFieldString','CUST46FV'),
    'CUST52FV':('dbo.vCustomFieldDate','CUST52FV'),
    # CDM column names (uppercase)
    'TOTALLOANAMOUNT':('dbo.vLoan','TotalLoanAmount'),
    'INTERESTRATE':('dbo.vLoan','InterestRate'),
    'NUMBEROFPAYMENTS':('dbo.vLoan','NumberofPayments'),
    'LOANNUMBER':('dbo.vLoan','LoanNumber'),
    'AMORTIZATIONTYPEENUM':('dbo.vLoan','AmortizationTypeEnum'),
    'LOANTYPEENUM':('dbo.vLoan','LoanTypeEnum'),
    'LOANPROGRAM':('dbo.vLoan','LoanProgram'),
    'LOANCHANNELENUM':('dbo.vLoan','LoanChannelEnum'),
    'ORIGINATORNMLS':('dbo.vLoan','OriginatorNMLS'),
    'SYSTEMOFRECORD':('dbo.vLoan','SystemOfRecord'),
    'MODIFIEDDATE':('dbo.vLoan','ModifiedDate'),
    'DATEAPPLICATION':('dbo.vLoan','DateApplication'),
    'LOANFOLDER':('dbo.vLoan','LoanFolder'),
    'LOAN_LOANFOLDER':('dbo.vLoan','LoanFolder'),
    'LOANPURPOSEENUM':('dbo.vProperty','LoanPurposeEnum'),
    'LIENPRIORITY':('dbo.vLoanProductData','LienPriority'),
    'DOCTYPEENUM':('dbo.vLoanProductData','DocTypeEnum'),
    'BRANCHMANAGERNMLS':('dbo.vLoanProductData','BranchManagerNMLS'),
    'UWRISKASSESSTYPEENUM':('dbo.vTsum','UWRiskAssessTypeEnum'),
    'AUSRECOMMENDATION':('dbo.vTsum','AUSRecommendation'),
    'SUBJECTPROPERTYTYPEENUM':('dbo.vTsum','SubjectPropertyTypeEnum'),
    'OCCUPANCYSTATUSENUM':('dbo.vApplication','OccupancyStatusEnum'),
    'BROKERCITY':('dbo.vContactBrokerLender','BrokerCity'),
    'BROKERSTATE':('dbo.vContactBrokerLender','BrokerState'),
    'BROKERLENDERNAME':('dbo.vContactBrokerLender','BrokerLenderName'),
    'LENDERSTATELICENSENBR':('dbo.vContactBrokerLender','LenderStateLicenseNbr'),
    'LOANPROCESSORNAME':('dbo.vContactLoanProcessor','LoanProcessorName'),
    'EMAILLOANPROCESSOR':('dbo.vContactLoanProcessor','EmailLoanProcessor'),
    'PROCESSOR':('dbo.vContactLoanProcessor','Processor'),
    'LOANTEAMMEMBER_EMAIL_LOAN_PROCESSOR':('dbo.vContactLoanProcessor','EmailLoanProcessor'),
    'LOANTEAMMEMBER_NAME_LOAN_PROCESSOR':('dbo.vContactLoanProcessor','LoanProcessorName'),
    'LOANTEAMMEMBER_EMAIL_DISCLOSER':('dbo.vContactLoanDiscloser','EmailDiscloser'),
    'EMAILDISCLOSER':('dbo.vContactLoanDiscloser','EmailDiscloser'),
    'LOANTEAMMEMBER_EMAIL_CLOSER':('dbo.vContactLoanCloser','EmailCloser'),
    'EMAILCLOSER':('dbo.vContactLoanCloser','EmailCloser'),
    'LOANTEAMMEMBER_NAME_CLOSER':('dbo.vContactLoanCloser','CloserName'),
    'CLOSERNAME':('dbo.vContactLoanCloser','CloserName'),
    'LOANTEAMMEMBER_EMAIL_FUNDER':('dbo.vContactLoanFunder','EmailFunder'),
    'EMAILFUNDER':('dbo.vContactLoanFunder','EmailFunder'),
    'FUNDER':('dbo.vContactLoanFunder','Funder'),
    'LOANTEAMMEMBER_EMAIL_UNDERWRITER':('dbo.vContactUnderWriter','EmailUnderwriter'),
    'EMAILUNDERWRITER':('dbo.vContactUnderWriter','EmailUnderwriter'),
    'CERTIFYINGUNDERWRITER':('dbo.vContactUnderWriter','CertifyingUnderwriter'),
    'LOANTEAMMEMBER_EMAIL_LOAN_OFFICER':('dbo.vContactLoanOfficer','EmailLoanOfficer'),
    'EMAILLOANOFFICER':('dbo.vContactLoanOfficer','EmailLoanOfficer'),
    'LOANTEAMMEMBER_NAME_LOAN_OFFICER':('dbo.vContactLoanOfficer','LoanOfficerName'),
    'LOANOFFICERNAME':('dbo.vContactLoanOfficer','LoanOfficerName'),
    'CORRACCOUNTEXECUTIVEEMAIL':('dbo.vLoanAssociate','CorrAccountExecutiveEmail'),
    'CORRACCOUNTEXECUTIVENAME':('dbo.vLoanAssociate','CorrAccountExecutiveName'),
    'BUYSIDELOCKDATE':('dbo.vRateLock','BuySideLockDate'),
    'PURCHASEADVICECONFIRMEDDATE':('dbo.vRateLock','PurchaseAdviceConfirmedDate'),
    'NEXTEXPECTEDMILESTONE':('dbo.vVirtualFields','NextExpectedMilestone'),
    'UWAPPRVALMONITORINGFICO':('dbo.vVaLoanData','UWApprvalMonitoringFICO'),
    'DATEESCROWINSURANCEEXPIRE':('dbo.vFhaVaLoan','DateEscrowInsuranceExpire'),
    'DATEGFE':('dbo.vRegulationZ','DateGFE'),
    'SECTIONACTTYPE':('dbo.vHudLoanData','SectionActType'),
    'FHASTREAMLINETYPE':('dbo.vHudLoanData','FHAStreamlineType'),
    'CORRESPONDENTUNDERWRITINGDELEGATEDENUM':('dbo.vTPO','CorrespondentUnderwritingDelegatedEnum'),
    'DATETPOSUBMIT':('dbo.vTPO','DateTPOSubmit'),
    'BORRLAST':('dbo.vLoanBorrower','BorrLast'),
    'DATECREDITDENIALACTIONTAKEN':('dbo.vStatementCreditDenial','DateCreditDenialActionTaken'),
    # Milestone log — DRS values (Log_MS_*)
    'LOG_MS_DATE_APPROVAL':('dbo.vMilestoneLog','DateApprove'),
    'LOG_MS_DATE_CLEAR_TO_CLOSE':('dbo.vMilestoneLog','DateClearToClose'),
    'LOG_MS_DATE_COMPLETION':('dbo.vMilestoneLog','DateCompletion'),
    'LOG_MS_DATE_DECISION':('dbo.vMilestoneLog','DateDecision'),
    'LOG_MS_DATE_DOCS_OUT':('dbo.vMilestoneLog','DateDocsOut'),
    'LOG_MS_DATETIME_FUNDING':('dbo.vMilestoneLog','DateFund'),
    'LOG_MS_DATE_PURCHASED':('dbo.vMilestoneLog','DatePurchase'),
    'LOG_MS_DATE_SHIPPING':('dbo.vMilestoneLog','DateShip'),
    'LOG_MS_DATETIME_SUBMITTAL':('dbo.vMilestoneLog','DateSubmittal'),
    'LOG_MS_DATE_PROCESSING':('dbo.vMilestoneLog','DateProcess'),
    'LOG_MS_DATE_RECONCILED':('dbo.vMilestoneLog','ReconciledDate'),
    'LOG_MS_DATETIME_APPROVAL':('dbo.vMilestoneLog','TimeApprovalDate'),
    'LOG_MS_STAGE':('dbo.vVirtualFields','NextExpectedMilestone'),
    # Milestone CDM column names
    'DATEAPPROVE':('dbo.vMilestoneLog','DateApprove'),
    'DATECLEARTOCLCLOSE':('dbo.vMilestoneLog','DateClearToClose'),
    'DATECOMPLETION':('dbo.vMilestoneLog','DateCompletion'),
    'DATEDECISION':('dbo.vMilestoneLog','DateDecision'),
    'DATEDOCSOUT':('dbo.vMilestoneLog','DateDocsOut'),
    'DATEFUND':('dbo.vMilestoneLog','DateFund'),
    'DATEPURCHASE':('dbo.vMilestoneLog','DatePurchase'),
    'DATEPROCESS':('dbo.vMilestoneLog','DateProcess'),
    'DATESUBMITTAL':('dbo.vMilestoneLog','DateSubmittal'),
    'DATESHIP':('dbo.vMilestoneLog','DateShip'),
    'RECONCILEDDATE':('dbo.vMilestoneLog','ReconciledDate'),
    'TIMEAPPROVALDATE':('dbo.vMilestoneLog','TimeApprovalDate'),
    'TIMECLEARTOCLOSEDATE':('dbo.vMilestoneLog','TimeClearToCloseDate'),
    'APPROVALSTATUS':('dbo.vMilestoneLog','ApprovalStatus'),
    'CLEARTOCLOSESTATUS':('dbo.vMilestoneLog','ClearToCloseStatus'),
    'DECISIONSTATUS':('dbo.vMilestoneLog','DecisionStatus'),
    'DOCSOUTSTATUS':('dbo.vMilestoneLog','DocsOutStatus'),
    'FUNDINGSTATUS':('dbo.vMilestoneLog','FundingStatus'),
    'SUBMITTALSTATUS':('dbo.vMilestoneLog','SubmittalStatus'),
}

# ── Out of Scope field sets ───────────────────────────────────────────────────
CALLIDUS_OOS = {
    'CX_BRANCHNMLS','CX_CENTRALIZED_1','CX_CMSPAIDOFFLN','CX_FISERVPRIORLOANNUMBER',
    'CX_MS_FCI_COUNT','CX_MS_READYFORDOCS_COUNT','CX_NY_CEMA','CX_PRODCENTER',
    'CX_PROJECTEDCLOSING2','CX_REFERRAL_EMAIL','CX_TX50F2','CX_TXHOMEEQUITY',
    'CX_UNFUNDDATE','CX_UW_2NDSIGDATE','CX_UW_2NDSIGNAME','CX_UW_CTCBY',
    'CX_UWQC_CNTCONDFORLP',
    'CUST01FV','CUST40FV','CUST46FV','CUST52FV',
    'CLOSERNAME','FUNDER','LOANPROCESSORNAME','PROCESSOR','CERTIFYINGUNDERWRITER',
    'LOANTEAMMEMBER_NAME_CLOSER','LOANTEAMMEMBER_NAME_LOAN_PROCESSOR',
    'F362','F984',
    'EMAILCLOSER','EMAILDISCLOSER','EMAILFUNDER','EMAILLOANPROCESSOR','EMAILUNDERWRITER',
    'LOANTEAMMEMBER_EMAIL_CLOSER','LOANTEAMMEMBER_EMAIL_DISCLOSER',
    'LOANTEAMMEMBER_EMAIL_FUNDER','LOANTEAMMEMBER_EMAIL_LOAN_PROCESSOR',
    'LOANTEAMMEMBER_EMAIL_UNDERWRITER',
    'CORRACCOUNTEXECUTIVEEMAIL','CORRACCOUNTEXECUTIVENAME',
    'LOANFOLDER','LOAN_LOANFOLDER','SYSTEMOFRECORD','MODIFIEDDATE',
    'BUYSIDELOCKDATE','PURCHASEADVICECONFIRMEDDATE','F2149','F2630',
    'SECTIONACTTYPE','DATEESCROWINSURANCEEXPIRE','DATEGFE','FHASTREAMLINETYPE',
    'F3000','F3089','F3152','F3659',
    'DATESHIP','LOG_MS_DATE_SHIPPING',
    'LOANPROCESSOR2EMAIL','LOANPROCESSOR3EMAIL',
    'LOANPROGRAM','LOANCHANNELENUM','F1401','F2626',
}

BANK_UNITED_OOS = {
    'LIFECAP','ARMMARGIN','PERIODICRATECAP','BALLOONTERM',
    'F247','F689','F697','F325',
    'MERSNUMBER','ESCROWCASENUMBER','TITLECASENUMBER','FUNDSWIRETO',
    'F1051','F186','F187','F2001',
    'AUSKEY','DOCTYPEENUM','MORNET_X4','MORNET_X67',
    'ESCROWCOABANBR','TITLECOABANBR',
    'VEND_X396','VEND_X397','VEND_X398','VEND_X399',
    'LOANCHANNELENUM','LOANPROGRAM','F2626','F1401',
}

_FHA_KEYS = {'SECTIONACTTYPE','DATEESCROWINSURANCEEXPIRE','DATEGFE','FHASTREAMLINETYPE',
             'F3000','F3089','F3152','F3659'}
_ARM_KEYS = {'LIFECAP','ARMMARGIN','PERIODICRATECAP','BALLOONTERM','F247','F689','F697','F325'}
_NULL_KEYS= {'LOANPROCESSOR2EMAIL','LOANPROCESSOR3EMAIL'}

def _oos_note(drs, cdm):
    for v in [drs, cdm]:
        if not v: continue
        k = v.upper()
        if k in _NULL_KEYS: return 'Hardcoded NULL in Callidus SQL (pGetEncompassData V76) — no Figure equivalent; Out of Scope'
        if k in _FHA_KEYS:  return 'FHA/HUD/GFE-specific field — not applicable to HELOC product; route to business stakeholders'
        if k in _ARM_KEYS:  return 'ARM-specific field — not applicable to HELOC product; route to business stakeholders'
    return 'CMH internal LOS workflow field — no Figure SFTP equivalent; route to business stakeholders'

def _is_oos(drs, cdm, tab_name):
    oos = CALLIDUS_OOS if 'Callidus' in tab_name else BANK_UNITED_OOS
    # special data entry error row
    if drs and 'decimal' in drs.lower():
        return True, 'Data entry error in source mapping spec — invalid field identifier; verify with business stakeholders'
    for v in [drs, cdm]:
        if v and v.upper() in oos:
            return True, _oos_note(drs, cdm)
    return False, ''

def _get_cdm(drs, cdm, tab_name):
    if 'Callidus' not in tab_name:
        return 'CDM', 'TBD — V3 (pending Bank United SQL)', 'TBD — V3'
    for v in [drs, cdm]:
        if v:
            hit = CDM_MAP.get(v.upper())
            if hit:
                return 'CDM', hit[0], hit[1]
    return 'CDM', 'TBD', 'TBD'

FIGURE_ORDER = [
    'warehouse-funding-queue','application-event','funding-transaction',
    'hmda','mcr','liability-payoff-reconciliation','dscr-application',
    'dscr-funding-transaction','originator-broker','user-identity',
    'loan-officer-license','task-outcome','complaint','waitlist'
]

# ── loaders ───────────────────────────────────────────────────────────────────
def load_doc1(path):
    wb = load_workbook(path)
    tabs = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            drs = str(r[0]).strip() if r[0] not in (None,'') else ''
            cdm = str(r[1]).strip() if r[1] not in (None,'') else ''
            if drs or cdm:
                rows.append({'drs': drs, 'cdm': cdm})
        tabs[name] = rows
    return tabs

def load_doc2(path):
    print("  Reading 45K rows (may take ~60s)...")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb['mssql_Field_Mapping_nopii']
    by_fid, by_fname, by_col = {}, {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        fid   = str(row[0]).strip() if row[0] else ''
        fname = str(row[1]).strip() if row[1] else ''
        dtype = str(row[2]).strip() if row[2] else ''
        desc  = str(row[6]).strip() if row[6] else ''
        rec   = {'FIELDID':fid,'FIELDNAME':fname,'DATATYPE':dtype,'DESCRIPTION':desc}
        if fid:  by_fid[fid.upper()] = rec
        if fname: by_fname[fname.lower()] = rec
        col = str(row[5]).strip().lower() if row[5] else ''
        if col: by_col[col] = rec
    wb.close()
    return by_fid, by_fname, by_col

def load_figure(base):
    schemas, lookup = {}, {}
    for fn in os.listdir(base):
        if not fn.endswith('.csv'): continue
        name = fn[:-4]
        fields = []
        with open(os.path.join(base, fn), encoding='utf-8') as f:
            for row in csv.DictReader(f):
                fn2 = row.get('Field Name','').strip()
                if fn2:
                    fields.append({'field_name':fn2,'datatype':row.get('Datatype','').strip(),
                                   'definition':row.get('Definition','').strip()})
        schemas[name] = fields
        lookup[name]  = {f['field_name']:f for f in fields}
    return schemas, lookup

# ── CMH leg-1 matching ────────────────────────────────────────────────────────
def norm_fid(fid):
    if re.match(r'^[Ff]\d', fid): return fid[1:].upper()
    return fid.upper()

def match_cmh(drs, cdm, by_fid, by_fname, by_col):
    rec = None; conf = 'Needs Review'; note = ''
    if drs:
        n = norm_fid(drs)
        if n in by_fid:
            rec=by_fid[n]; conf='Confirmed'; note=f'Pass 1 exact FID {drs}->{n}'
        elif drs.upper() in by_fid:
            rec=by_fid[drs.upper()]; conf='Confirmed'; note=f'Pass 1 direct {drs}'
        else:
            nf = re.sub(r'[\s\-_]','',n)
            for k,v in by_fid.items():
                if re.sub(r'[\s\-_]','',k)==nf:
                    rec=v; conf='High'; note=f'Pass 2 fuzzy {drs}->{k}'; break
    if not rec and cdm:
        cl = cdm.lower()
        if cl in by_fname:
            rec=by_fname[cl]; conf='Probable'; note=f'Pass 3 name match {cdm}'
        elif cl in by_col:
            rec=by_col[cl]; conf='Probable'; note=f'Pass 3 col match {cdm}'
        else:
            ct = tokenize(cdm) - {'enum','type','status','none','nan'}
            best,br = 0,None
            for fn,rv in by_fname.items():
                ft = tokenize(fn)
                if ft:
                    s = len(ct&ft)/max(len(ct),len(ft))
                    if s>best: best=s; br=rv
            if best>=0.7 and br:
                rec=br; conf='Probable'; note=f'Pass 3 token {best:.0%} on {cdm}'
    if not rec:
        note = f'Pass 4 no match FID={drs} Name={cdm}'
    return {
        'CMH_FieldName':    rec['FIELDNAME']   if rec else '',
        'CMH_DataType':     rec['DATATYPE']    if rec else '',
        'CMH_Description':  rec['DESCRIPTION'] if rec else '',
        'CMH_Match_Confidence': conf,
        'CMH_Match_Notes':  note,
    }

# ── Figure leg-2 matching ─────────────────────────────────────────────────────
def fig_confidence(schema, priority, notes):
    if not schema: return 'No Figure Match'
    n = notes or ''
    if 'HELOC FLAG' in n or 'approximation' in n.lower(): return 'High'
    if 'Algorithmic' in n: return 'Probable'
    if priority=='PRIMARY' and schema=='warehouse-funding-queue': return 'Confirmed'
    return 'High'

def algo_match(search, schemas):
    toks = tokenize(search) - {'enum','type','status','none','nan','indicator','flag'}
    if not toks: return []
    hits = []
    for sname in FIGURE_ORDER:
        if sname not in schemas: continue
        for f in schemas[sname]:
            ft = tokenize(f['field_name'])
            if not ft: continue
            s = len(toks&ft)/max(len(toks),len(ft))
            if s>=0.5: hits.append((s,sname,f))
    hits.sort(key=lambda x:(-x[0], FIGURE_ORDER.index(x[1]) if x[1] in FIGURE_ORDER else 99))
    seen=set(); out=[]
    for _,sn,f in hits:
        if sn not in seen: seen.add(sn); out.append((sn,f))
    return out[:2]

def build_rows(drs, cdm, cmh, schemas, lookup, tab_name=''):
    keys = [k for k in [drs, cdm] if k]
    fig_matches = None
    for k in keys:
        if k in M: fig_matches=M[k]; break

    cdm_db, cdm_view, cdm_col = _get_cdm(drs, cdm, tab_name)

    BASE_ROW = {
        'DRSColumnName': drs, 'CDMViewColumnName': cdm,
        'CMH_FieldName': cmh['CMH_FieldName'], 'CMH_DataType': cmh['CMH_DataType'],
        'CMH_Description': cmh['CMH_Description'],
        'CMH_Match_Confidence': cmh['CMH_Match_Confidence'],
        'CMH_Match_Notes': cmh['CMH_Match_Notes'],
        'CDM_Database': cdm_db,
        'CDM_View': cdm_view,
        'CDM_Column': cdm_col,
        'Schema_Pending': 'Yes',
    }

    def make(schema, field_name, priority, notes, status):
        fig_f = lookup.get(schema,{}).get(field_name) if schema else None
        r = dict(BASE_ROW)
        r.update({
            'Figure_Field_Name':       field_name or '',
            'Figure_DataType':         fig_f['datatype']   if fig_f else '',
            'Figure_Definition':       fig_f['definition'] if fig_f else '',
            'Figure_Source_File':      schema or '',
            'Figure_Priority':         priority or '',
            'Figure_Match_Confidence': fig_confidence(schema, priority, notes),
            'Figure_Match_Notes':      notes or '',
            'Overall_Status':          status,
        })
        return r

    if fig_matches is None:
        search = cmh['CMH_FieldName'] or cdm or drs or ''
        algo = algo_match(search, schemas)
        if algo:
            fig_matches = [(sn, f['field_name'], 'PRIMARY' if i==0 else 'SECONDARY', 'Algorithmic match')
                           for i,(sn,f) in enumerate(algo)]
        else:
            fig_matches = [None]

    rows_out = []
    if not fig_matches or fig_matches == [None]:
        rows_out.append(make(None,'',None,'No equivalent found in any Figure schema','No Figure Match'))
    else:
        for m in fig_matches:
            if m is None:
                rows_out.append(make(None,'',None,'No equivalent found in any Figure schema','No Figure Match'))
                break
            sn, fn, pri, notes = m
            rows_out.append(make(sn, fn, pri, notes, 'Mapped'))

    # Out of Scope override: applies only when ALL rows are No Figure Match
    if all(r['Overall_Status']=='No Figure Match' for r in rows_out):
        oos, oos_note = _is_oos(drs, cdm, tab_name)
        if oos:
            for r in rows_out:
                r['Overall_Status']  = 'Out of Scope'
                r['Figure_Match_Notes'] = oos_note

    return rows_out

# ── Excel output ──────────────────────────────────────────────────────────────
HEADERS = [
    'DRSColumnName','CDMViewColumnName',
    'CMH_FieldName','CMH_DataType','CMH_Description',
    'CMH_Match_Confidence','CMH_Match_Notes',
    'CDM_Database','CDM_View','CDM_Column',
    'Figure_Field_Name','Figure_DataType','Figure_Definition',
    'Figure_Source_File','Figure_Priority',
    'Figure_Match_Confidence','Figure_Match_Notes',
    'Schema_Pending','Overall_Status'
]
#              1   2   3   4   5   6   7  8   9  10  11 12 13  14 15  16 17  18  19
COL_W = [22, 25, 28, 14, 42, 22, 38, 12, 22, 25, 30, 16, 42, 30, 14, 24, 42, 15, 20]

FILLS = {
    'Confirmed':      PatternFill('solid', fgColor='C6EFCE'),
    'High':           PatternFill('solid', fgColor='FFEB9C'),
    'Probable':       PatternFill('solid', fgColor='FCE4D6'),
    'Needs Review':   PatternFill('solid', fgColor='FFC7CE'),
    'No Figure Match':PatternFill('solid', fgColor='D9D9D9'),
    'Out of Scope':   PatternFill('solid', fgColor='BDD7EE'),
    'Mapped':         PatternFill('solid', fgColor='C6EFCE'),
}
HDR_FILL = PatternFill('solid', fgColor='1F4E79')
HDR_FONT = Font(bold=True, color='FFFFFF')

def write_excel(path, tab_data):
    wb = Workbook()
    first = True
    for tab_name, rows in tab_data.items():
        ws = wb.active if first else wb.create_sheet()
        ws.title = tab_name; first = False
        for c,h in enumerate(HEADERS,1):
            cell = ws.cell(1,c,h)
            cell.font=HDR_FONT; cell.fill=HDR_FILL
            cell.alignment=Alignment(horizontal='center',wrap_text=True)
        ws.row_dimensions[1].height = 28
        for ri, row in enumerate(rows,2):
            for ci,h in enumerate(HEADERS,1):
                v = row.get(h,'')
                cell = ws.cell(ri,ci,v)
                cell.alignment=Alignment(wrap_text=True,vertical='top')
                if h in ('CMH_Match_Confidence','Figure_Match_Confidence','Overall_Status'):
                    fill = FILLS.get(v)
                    if fill: cell.fill=fill
        for ci,w in enumerate(COL_W,1):
            ws.column_dimensions[get_column_letter(ci)].width=w
        ws.freeze_panes='A2'
    wb.save(path)
    print(f"Saved -> {path}")

# ── main ──────────────────────────────────────────────────────────────────────
def main():
    print("Step 1/5  Loading Doc 1...")
    doc1 = load_doc1(DOC1_PATH)
    for t,r in doc1.items(): print(f"  {t}: {len(r)} rows")

    print("Step 2/5  Loading Doc 2 (45K rows)...")
    by_fid, by_fname, by_col = load_doc2(DOC2_PATH)
    print(f"  FID index: {len(by_fid)}")

    print("Step 3/5  Loading Figure schemas...")
    fig_schemas, fig_lookup = load_figure(FIG_BASE)
    print(f"  {len(fig_schemas)} schemas loaded")

    print("Step 4/5  Running matching...")
    tab_out = {}
    for tab_name, rows in doc1.items():
        out = []
        for row in rows:
            drs, cdm = row['drs'], row['cdm']
            if not drs and not cdm: continue
            cmh = match_cmh(drs, cdm, by_fid, by_fname, by_col)
            out.extend(build_rows(drs, cdm, cmh, fig_schemas, fig_lookup, tab_name))
        tab_out[tab_name] = out
        oos_count  = sum(1 for r in out if r['Overall_Status']=='Out of Scope')
        map_count  = sum(1 for r in out if r['Overall_Status']=='Mapped')
        nfm_count  = sum(1 for r in out if r['Overall_Status']=='No Figure Match')
        print(f"  {tab_name}: {len(rows)} in -> {len(out)} out  "
              f"(Mapped={map_count} / OutOfScope={oos_count} / NoFigMatch={nfm_count})")

    print("Step 5/5  Writing Excel...")
    write_excel(OUT_PATH, tab_out)
    print("DONE")

if __name__ == '__main__':
    main()
